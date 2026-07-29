# -*- coding: utf-8 -*-
"""
Cruza el Excel de la empresa (productos_organizados_2.xlsx) con los productos
reales de WooCommerce (wc-products-images.json) y agrega columnas con los
links de imagen (media de WordPress) para verificacion manual foto por foto.
"""
import json
import re
import unicodedata
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

SRC_XLSX = r"C:\Users\jonat\Downloads\productos_organizados_2.xlsx"
OUT_XLSX = r"C:\Users\jonat\Downloads\productos_organizados_2_con_links_imagenes.xlsx"
WC_JSON = "wc-products-images.json"

CATEGORY_PREFIXES = [
    "Foliar", "Fertilizante", "Fungicida", "Insecticida", "Herbicida",
    "Acaricida", "Molusquicida", "Adherente", "Rodenticida", "Herramienta",
    "Semilla", "Repuesto", "Articulo", "Artículo",
]


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def remove_category_prefix(s):
    for p in CATEGORY_PREFIXES:
        if s.startswith(p + " "):
            s = s[len(p) + 1:].strip()
    if s.startswith("Victorinox "):
        s = s[len("Victorinox "):]
    return s


def canonicalize(s):
    if not s:
        return ""
    s = strip_accents(s).lower()
    return re.sub(r"[^a-z0-9]", "", s)


def tokenize(s):
    if not s:
        return []
    s = strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]", " ", s)
    return [t for t in s.split() if len(t) >= 3]


def load_wc_products():
    with open(WC_JSON, encoding="utf-8") as f:
        products = json.load(f)
    for p in products:
        base = remove_category_prefix(p["name"])
        p["_canon"] = canonicalize(base)
        p["_tokens"] = tokenize(base)
    return products


def find_best_match(query_text, products):
    q_canon = canonicalize(remove_category_prefix(query_text))
    if not q_canon:
        return None, None

    for p in products:
        if p["_canon"] == q_canon:
            return p, "exacto"

    for p in products:
        if len(q_canon) >= 6 and (q_canon in p["_canon"] or p["_canon"] in q_canon):
            return p, "substring"

    q_tokens = tokenize(remove_category_prefix(query_text))
    if q_tokens:
        best, best_score = None, 0.0
        for p in products:
            if not p["_tokens"]:
                continue
            matches = sum(1 for t in q_tokens if t in p["_tokens"])
            score = matches / len(q_tokens)
            if score > best_score and score >= 0.6:
                best_score = score
                best = p
        if best:
            return best, f"tokens({best_score:.2f})"

    return None, None


REVIEW_WC_IDS = {
    334: "El nombre no coincide con la foto (foto parece de otro producto/línea)",
    320: "El nombre no coincide con la foto (foto parece de otro producto/línea)",
    300: "Nombre de archivo no concluyente, verificar visualmente",
    281: "La foto es de un formato distinto (pellet vs bloque)",
    244: "La foto corresponde a otro ingrediente activo (bifentrin, no cyperkill)",
    194: "La foto parece ser de otro fertilizante (fosfato diamónico)",
    182: "El nombre del archivo no corresponde al producto, verificar",
    168: "Foto genérica reutilizada de otra línea (Agri Producción), verificar si es la correcta",
    167: "Foto genérica reutilizada de otra línea (Agri Producción), verificar si es la correcta",
    166: "Foto genérica reutilizada de otra línea (Agri Producción), verificar si es la correcta",
    154: "La foto parece ser de otro producto (Veca Move)",
}


def main():
    wc_products = load_wc_products()

    wb = openpyxl.load_workbook(SRC_XLSX)
    main_ws = wb["TODOS LOS PRODUCTOS"]

    # header map for main sheet
    headers = [c.value for c in main_ws[1]]
    col_codigo = headers.index("CODIGO")
    col_producto = headers.index("PRODUCTO")
    col_formato = headers.index("FORMATO") if "FORMATO" in headers else None

    match_by_codigo = {}
    stats = {"total": 0, "matched": 0, "with_image": 0, "with_2_images": 0, "unmatched": 0}

    for row in main_ws.iter_rows(min_row=2):
        codigo = row[col_codigo].value
        producto = row[col_producto].value
        formato = row[col_formato].value if col_formato is not None else None
        if not codigo or not producto:
            continue
        stats["total"] += 1

        query_text = producto if not formato else f"{producto} {formato}"
        match, match_type = find_best_match(query_text, wc_products)

        images = match["images"] if match else []
        wc_id = match["id"] if match else None
        result = {
            "wc_name": match["name"] if match else "",
            "wc_id": wc_id,
            "match_type": match_type or "",
            "img1": images[0]["src"] if len(images) > 0 else "",
            "img2": images[1]["src"] if len(images) > 1 else "",
            "observacion": REVIEW_WC_IDS.get(wc_id, ""),
        }
        match_by_codigo[str(codigo).strip()] = result

        if match:
            stats["matched"] += 1
        else:
            stats["unmatched"] += 1
        if len(images) >= 1:
            stats["with_image"] += 1
        if len(images) >= 2:
            stats["with_2_images"] += 1

    # Style reference from an existing header cell
    header_font = copy(main_ws.cell(row=1, column=1).font)
    header_fill = copy(main_ws.cell(row=1, column=1).fill)
    header_align = copy(main_ws.cell(row=1, column=1).alignment)

    NEW_HEADERS = ["PRODUCTO WC (coincidencia)", "LINK IMAGEN 1", "LINK IMAGEN 2", "OBSERVACION"]

    for ws in wb.worksheets:
        sheet_headers = [c.value for c in ws[1]]
        if "CODIGO" not in sheet_headers:
            continue
        codigo_col = sheet_headers.index("CODIGO") + 1  # 1-indexed
        start_col = ws.max_column + 1

        for i, h in enumerate(NEW_HEADERS):
            cell = ws.cell(row=1, column=start_col + i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx in range(2, ws.max_row + 1):
            codigo_val = ws.cell(row=row_idx, column=codigo_col).value
            if not codigo_val:
                continue
            info = match_by_codigo.get(str(codigo_val).strip())
            if not info:
                continue
            ws.cell(row=row_idx, column=start_col, value=info["wc_name"])
            ws.cell(row=row_idx, column=start_col + 1, value=info["img1"])
            ws.cell(row=row_idx, column=start_col + 2, value=info["img2"])
            obs_cell = ws.cell(row=row_idx, column=start_col + 3, value=info["observacion"])
            if info["observacion"]:
                fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                for c in range(1, start_col + 4):
                    ws.cell(row=row_idx, column=c).fill = fill

        # widen new columns a bit
        for i in range(len(NEW_HEADERS)):
            col_letter = ws.cell(row=1, column=start_col + i).column_letter
            ws.column_dimensions[col_letter].width = 45

    wb.save(OUT_XLSX)

    print("Guardado:", OUT_XLSX)
    print("Resumen (sobre 'TODOS LOS PRODUCTOS'):")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
